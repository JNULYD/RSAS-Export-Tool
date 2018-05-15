import os
import re
import sys
import html
import time
import zipfile
import datetime
import openpyxl
import Coredata.load


class Vul_re(object):
    def __init__(self):
        super(Vul_re, self).__init__()
        self.content_re = '<python>python<python>(.*?)<python>python</python>'
        self.vul_list_re = '<python>host<python>.*?<td valign="top".*?<th width="120">IP地址</th>.*?<td>(.*?)</td>.*?</td>.*?<python>host</python>.*?<python>vul_list<python>(.*?)<python>vul_list</python>'
        self.vul_detail_re = '<python>vul_detail<python>(.*?)<python>vul_detail</python>'
        self.vul_details_re = '<python>vul_details<python>(.*?)<python>vul_details</python>'

        self.danger_re = '<span class="level_danger_(.*?)".*?(table_\d_\d+).*?>(.*?)</span>'
        self.title_re = '<python>title<python>(.*?)<python>title</python>'
        self.other_re = '<td class="vul_port">(.*?)</td>.*?<td>(.*?)</td>.*?<td>(.*?)</td>.*?<td>.*?<ul>(.*?)</ul>'
        self.hostname_re = '<python>host<python>.*?<td valign="top".*?<th width="120">IP地址</th>.*?<th>主机名</th>.*?<td>(.*?)</td>.*?</td>.*?<python>host</python>'
        self.systemname_re = '<python>host<python>.*?<td valign="top".*?<th width="120">IP地址</th>.*?<th>主机名</th>.*?<th width="120">操作系统</th>.*?<td>(.*?)</td>.*?</td>.*?<python>host</python>'

class Vul_content(object):
    def __init__(self,vul_re):
        super(Vul_content, self).__init__()
        self.vul_list_content = re.findall(vul_re.vul_list_re,htmlcont,re.S|re.M)
        self.vul_detail_content = re.findall(vul_re.vul_detail_re,htmlcont,re.S|re.M)


class Solve_re(object):
    def __init__(self, danger):
        super(Solve_re, self).__init__()
        self.solve_re = '<tr class="solution.*?%s.*?<th width="100">解决办法</th>.*?<td>(.*?)</td>' % danger
        self.describe_re = '<tr class="solution.*?%s.*?<th width="100">详细描述</th>.*?<td>(.*?)</td>' % danger
        self.cve_re = '<tr class="solution.*?%s.*?<th width="100">CVE编号</th>.*?<td><a target=.*?>(.*?)</a>.*?</td>' % danger

class Host(object):
    def __init__(self, vul_re):
        super(Host, self).__init__()
        self.hostname = re.findall(vul_re.hostname_re,htmlcont,re.S|re.M)
        self.systemname = re.findall(vul_re.systemname_re,htmlcont,re.S|re.M)

class Other(object):
    def __init__(self, vul_re, all_vuln_list):
        super(Other, self).__init__()
        self.all_other = re.findall(vul_re.other_re,all_vuln_list,re.S|re.M)

class Danger(object):
    def __init__(self, vul_re, other):
        super(Danger, self).__init__()
        #获取漏洞名称、与之相关的漏洞table标签
        self.danger_coneent = re.findall(vul_re.danger_re,other,re.S|re.M)

class Solve(object):
    def __init__(self, solve, all_vul_detail):
        super(Solve, self).__init__()
        self.solve_plumb = re.findall(solve.solve_re,all_vul_detail,re.S|re.M)
        self.describe_plumb = re.findall(solve.describe_re,all_vul_detail,re.S|re.M)
        self.cve_plumb = re.findall(solve.cve_re,all_vul_detail,re.S|re.M)

def main():
    folder_name = str(sys.argv[1])
    Coredata.load.load_date()

    content = open('./Coredata/database.mdb','r',encoding='utf-8')
    all_cont = content.read()
    content.close()
    print('\n数据提取完成，正在生成漏洞跟踪表...')
    wb = openpyxl.Workbook()
    ws = wb.active


    vul_re = Vul_re()
    #正则：<python>python<python>(.*?)<python>python</python>
    all_list_content = re.findall(vul_re.content_re,all_cont,re.S|re.M)

    global htmlcont
    for htmlcont in all_list_content:
        sheet_name =  re.findall(vul_re.title_re,htmlcont,re.S|re.M)
        for sheet in sheet_name:
            print('正在导出 %s,请稍后...'%sheet)
            ws = wb.create_sheet(sheet,0)
            ws['A1'] = '序号'
            ws['B1'] = '主机名'
            ws['C1'] = '操作系统'
            ws['D1'] = 'IP地址'
            ws['E1'] = '发现风险点描述'
            ws['F1'] = '风险分类'
            ws['G1'] = '风险等级'
            ws['H1'] = '整改建议'
            ws['I1'] = '漏洞概述'
            ws['J1'] = '漏洞CVE编号'
            ws['K1'] = '漏洞对应端口'
            ws['L1'] = '漏洞对应协议'
            ws['M1'] = '漏洞对应服务'

        vul_content = Vul_content(vul_re)
        hostname = Host(vul_re).hostname
        if hostname:
            pass
        else:
            hostname = ['']
        systemname = Host(vul_re).systemname
        if systemname:
            pass
        else:
            systemname = ['']
        x = 0
        for all_vul_list in vul_content.vul_list_content:
            for other in Other(vul_re,all_vul_list[1]).all_other:
                for danger in Danger(vul_re,other[3]).danger_coneent:
                    x += 1
                    for all_vul_detail in vul_content.vul_detail_content:
                        vul_details_content = re.findall(vul_re.vul_details_re,all_vul_detail,re.S|re.M)
                        for all_vul_details in vul_details_content:
                            vul_detail = Solve(Solve_re(danger[1]),all_vul_details)
                            for describe,solve in zip(vul_detail.describe_plumb,vul_detail.solve_plumb):
                                cve = vul_detail.cve_plumb
                                if cve:
                                    pass
                                else:
                                    cve = ['此漏洞暂无CVE漏洞编号。']
                    ws.append([x,hostname[0],systemname[0],all_vul_list[0],danger[2],'漏洞',danger[0].replace('low','低').replace('middle','中').replace('high','高'),html.unescape(re.sub('<br/>','',solve)).replace(' ','').strip('\n'),html.unescape(re.sub('<br/>','',describe)).replace(' ','').strip('\n'),cve[0],other[0],other[1],other[2]])
    wb.save(folder_name+'/漏洞跟踪表.xlsx')
    print('\n恭喜！漏洞跟踪表导出完成，保存在 %s 目录下。'%folder_name)

if __name__ == '__main__':
    main()
    Coredata.load.clean_date()